import sys
import os
import shutil
from openpyxl import load_workbook


def excel_json(fp, idx, data):
    fp.write('{\n')
    for row in data.iter_rows(min_row = 2):
        key = row[0].value
        val = ''
        if row[idx].value is not None:
            val = row[idx].value.replace('\n', " <br> ")
            val = val.replace('\\', '\\\\')
        str = f'    \"{key}\": \"{val}\",\n'
        fp.write(str)
    fp.write('}\n')


if len(sys.argv) != 2:
    print("Insufficient arguments")
    sys.exit()

print('==== process start ====\n')
file_path = sys.argv[1]
print("File path : " + file_path)

# directory_path = ['de', 'en', 'en-AU', 'en-GB', 'en-US', 'fr', 'it', 'ja']
directory_path = ['en', 'ja', 'it', 'de', 'fr']
root_path = './'
locale_path = root_path + 'locales/'
target_file_name = 'translation.json'

wb = load_workbook(file_path, read_only=True)
data = wb.active

idx = 2
for path in directory_path:
    target_path = locale_path + path
    os.makedirs(target_path, exist_ok=True)
    fp = open(target_path + '/' + target_file_name, 'w')
    excel_json(fp, idx, data)
    fp.close()
    idx = idx + 1


os.makedirs(locale_path + 'en-AU', exist_ok=True)
os.makedirs(locale_path + 'en-GB', exist_ok=True)
os.makedirs(locale_path + 'en-US', exist_ok=True)
shutil.copy(locale_path + 'en' + '/' + target_file_name, locale_path + 'en-AU' + '/' + target_file_name)
shutil.copy(locale_path + 'en' + '/' + target_file_name, locale_path + 'en-GB' + '/' + target_file_name)
shutil.copy(locale_path + 'en' + '/' + target_file_name, locale_path + 'en-US' + '/' + target_file_name)

print('==== process end ====\n')
